import os
from datetime import datetime
import logging

from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import Column, ForeignKey, Integer, String, Boolean, delete, update
from sqlalchemy.orm import sessionmaker
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from connection import engine, ROOT

#Configuration for logging
logging.basicConfig(filename = ROOT / "log.log", \
                    format="%(asctime)s %(levelname)-10s %(message)s", \
                    level = logging.DEBUG)

Base = declarative_base()

def create_metadata():
    Base.metadata.create_all(engine)

Session = sessionmaker(bind = engine)
session = Session()


class Employee(Base):

    """The class represents an employee. Through this class, employees can be added 
       or deleted, as well as some data related to employees can be modified."""

    __tablename__ = 'Employees'
    num_id = Column(Integer, primary_key=True)
    full_name = Column(String)
    working_hours  = Column(Integer)
    available_lines = Column(Integer, default = 0)
    
    def __init__(self, full_name, working_hours):
        self.full_name = full_name
        self.working_hours = working_hours    

    def add_employee(self):

        """Through this method, an employee can be added to the database. No arguments needed."""
        
        try:
            self.full_name = input("Enter the full name: ").capitalize().strip()
            self.working_hours = float(input("Enter the working hours per day: "))
        except ValueError:
            print("There is a writing error. Please try to enter the data again carefully.")
            logging.error(f"New employee: The data was not entered correctly.")
            return
        new_employee = Employee(full_name = self.full_name,
                                working_hours = self.working_hours)
        new_employee.available_lines = self.working_hours * 10
        session.add(new_employee)
        session.commit()
        logging.info(f"New employee added. Full name: {new_employee.full_name} ")

    def show_employees(self):

        """A list of all employees will be displayed."""

        list = session.query(Employee).all()
        print("=" * 30)
        print ("The list with employees.")
        print("=" * 30)
        for row in list:
            print(f"ID: {row.num_id: <3} Name: {row.full_name: <20} Working hours per day: {row.working_hours}")
    
    def delete_item(self):

        """Through this method, an employee can be deleted from the database. 
           No arguments needed."""

        user_input = int(input("To delete an employee, please enter the ID: "))
        try:
            emp_name = session.query(Employee).filter(Employee.num_id == user_input).first().full_name
        except AttributeError:
            print("The user ID doesn`t exist. Please try again!")
            logging.error("No User ID found.")
            return

        else:
            session.execute(delete(Employee).where(Employee.num_id == user_input))
            session.commit()
            print(f"Employee {emp_name} has been successfully deleted.")
        logging.info(f"An employee was deleted. Last name: {emp_name}")

    def recharge_available_lines(self):

        """Through this method, the employees' available hours will be reallocated. 
           Considering that each line in the task requires 6 minutes of processing, 
           each available hour represents 10 lines in the task."""

        employees = session.query(Employee).order_by(Employee.available_lines).all()
        for employee in employees:
            hours = employee.working_hours * 10
            session.execute(update(Employee).where(Employee.num_id == employee.num_id).values(available_lines = hours))
        session.commit() 
        print ("the employees' available hours/lines were recharged.")
        logging.info("The employees' available hours/lines were recharged.")
        return

class Task(Base):

    """The class represents a task. Through this class, tasks can be added or deleted."""

    __tablename__ = 'Tasks'
    num_id = Column(Integer, primary_key=True)

    date = Column(String)
    hour = Column(String)
    name = Column(String)
    lines  = Column(Integer)
    available_lines = Column(Integer, default=lambda self: self.lines)
    processed = Column(Boolean, default = False)
    

    def __init__(self, date, hour, name, lines):
        self.date = date
        self.hour = hour
        self.name = name
        self.lines = lines
        self.available_lines = lines


    def completed_task(self):

        """Through this method, a certain task will be signed as processed. 
           The basic criteria are the available lines. When all lines were assigned the 
           'processed' attribute will be True"""

        tasks = session.query(Task).filter_by(available_lines = 0).all()
        for task in tasks:
            session.execute(update(Task).where(Task.available_lines == 0).values(processed = True))
            logging.info(f"The completed tasks {task.name} were signed as processed.")
        session.commit()


    def add_task(self):

        """Through this method, a certain task will be deleted."""

        try:
            self.name = str(input("Enter the task`s name: ")).title().strip()
            self.lines = int(input("Enter the number of line: "))
        except ValueError:
            print("There is a writing error. Please try to enter the data again carefully.")
            logging.error(f"New task: The data was not entered correctly.")
            return

        now = datetime.now()
        new_task = Task(date = now.strftime("%d/%m/%Y"),
                        hour = now.strftime("%H:%M:%S"),
                        name = self.name,
                        lines = self.lines,)
        session.add(new_task)
        session.commit()
        print(f"The task {new_task.name} was added.")
        logging.info(f"The task {new_task.name} was added.")


    def show_tasks(self):

        """A list of all tasks will be displayed."""

        list = session.query(Task).all()
        print("=" * 30)
        print ("The list with tasks.")
        print("=" * 30)
        for row in list:
            print (F"ID: {row.num_id: <4} Date: {row.date: <13} Hour: {row.hour: <12} Name: {row.name: <13} Nr. of lines {row.lines}")

    def delete_item(self):

        """Through this method, a certain task will be deleted. 
           To be able to delete the task, you need its id number"""

        try:
            user_input = int(input("To delete a task, please enter the ID: "))
            task_name = session.query(Task).filter(Task.num_id == user_input).first().name
        except (ValueError, AttributeError):
            print("No ID number introduced or could not be found.")
            logging.error("Delete item: No ID number introduced.")
            return
        else:
            session.execute(delete(Task).where(Task.num_id == user_input))
            session.commit()
            print(f"The task {task_name} was on purpose deleted.")
            logging.info(f"The task {task_name} was on purpose deleted.")
            

class Plan(Base):
     
    """This class is meant to create a plan object. It includes 
       methods for calculating the work schedule for a working day. 
       Also, through this class, writing to the Excel file will be performed."""

    __tablename__ = 'Plan'
    num_id = Column(Integer, primary_key=True)

    date        = Column(String)
    hour        = Column(String)
    task_name   = Column(String)
    users       = Column(String)
    lines       = Column(Integer)
    processed   = Column(Boolean, default = False)
    

    def __init__(self, date, hour, task_name, users, lines):
        self.date = date
        self.hour = hour
        self.task_name = task_name
        self.users = users
        self.lines = lines
        self.today = datetime.now().strftime("%d.%m.%Y")
        self.xlsx_name = f"Plan {self.today}.xlsx"
        self.file_path = os.path.join(ROOT, self.xlsx_name)

    def create_plan(self):

        """The method takes from another classes all necessary 
           data to calculate the time plan object."""
        
        while True:
            # create two lists of available tasks and employees
            tasks = session.query(Task).order_by(Task.processed.asc(), Task.date.asc(), Task.available_lines.asc()).all()
            employees = session.query(Employee).order_by(Employee.available_lines.desc()).all()

            # verifying if the tasks were full assigned or if 
            # the employees` availability is at maximum capacity
            if all(task.available_lines == 0 for task in tasks) or \
               all(employee.available_lines == 0 for employee in employees):
                break
            
            for task in tasks:
                for employee in employees:
                    if task.available_lines == 0 or employee.available_lines == 0:
                        continue
                    availability = min(task.available_lines, employee.available_lines)

                    new_plan = Plan(
                        date = self.today,
                        hour = task.hour,
                        task_name = task.name,
                        users = employee.full_name,
                        lines = availability)

                    task.available_lines -= availability
                    employee.available_lines -= availability

                    session.add(new_plan)
                    session.commit()
                    Task.completed_task(Task)
        print("The plan was created.")
        logging.info(f"The plan for {self.today} was created.")
        return

    def create_workbook(self):

        if os.path.exists(self.file_path):
            print("*" * 30)
            print("The plan has already been created. Please check your folder!")
            print("*" * 30)
            logging.warning("The xlsx file has already been created.")
            return
        else:
            try:
                wb = Workbook()
                wb.save(self.file_path)
                logging.info(f"The xlsx file was created. Name: {self.xlsx_name}")
            except OSError as err:
                print(err)
                logging.error(err)
    
    
    def update_plan(self, the_query_list):

        """This function is intended to update the created plan. The function should always 
        be used after writing the data into a xlsx file. The iteration through the plan, 
        more precisely the query variable, will be considered as a parameter."""

        for task in the_query_list:
            session.execute(update(Plan).where(Plan.processed == False).values(processed = True))
        session.commit()
        logging.info(f"The data was written. File name: {self.xlsx_name}")


    def write_plan(self):
        
        """This method is responsible for transferring the plan to an xlsx file."""

        plan_list = session.query(Plan).filter(Plan.processed == 0).order_by(Plan.date.asc(), Plan.hour.asc()).all()

        self.create_workbook()
        wb = load_workbook(self.file_path)
        ws = wb.active

        row_idx = 2
        for task in plan_list:
            col_idx = 1
            for attribute in [task.date, task.hour, task.task_name, task.users, task.lines]:
                ws.cell(row = row_idx, column = col_idx, value = attribute)
                col_idx += 1
            row_idx += 1
        
        ws.merge_cells('A1:E1')
        ws["A1"] = self.today
        title_left_cell = ws["A1"]
        title_left_cell.alignment = Alignment(horizontal = "center")
        title_left_cell.font = Font(b = True)
        fill_cell = PatternFill(fill_type="solid", fgColor="00FFFF99")
        title_left_cell.fill = fill_cell

        wb.save(self.file_path)
        self.update_plan(plan_list)
        print(f"The file Plan {self.today}.xlsx has been created.")
        logging.info(f"The {self.today} XLSX file has been updated.")